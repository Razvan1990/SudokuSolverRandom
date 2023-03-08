import os
import time

import pandas as pd
import openpyxl
from openpyxl.styles import Side, Border, PatternFill, Font
from constants import constants
from game_preparing.number_creator import CreateNumbers


class ExcelWriter:

    def __init__(self):
        self.creator = CreateNumbers()

    def check_if_file_closed(self):
        print("Please close the file in order to proceed with the solving of the puzzle. \n\
        Please confirm by typing an Y(y) if you have closed the file")
        x = input("")
        while x.upper() != "Y":
            print("We are still waiting....")
            x = input("")
        return x

    def write_to_file(self, list_excel):

        file_path = constants.FILE_PATH
        os.chdir(file_path)
        # change zeroes into blanks for Excel file
        for i in range(0, 9):
            for j in range(0, 9):
                if list_excel[i][j] == 0:
                    list_excel[i][j] = " "
        df_excel = pd.DataFrame(list_excel, index=constants.ROW_NAMES, columns=constants.COLUMN_NAMES)
        df_excel.to_excel(constants.NAME_FILE, sheet_name="initial game")
        wb = openpyxl.load_workbook(filename=constants.NAME_FILE)
        work_sheet = wb["initial game"]
        border_design = Side(style="thick", color="FF0000")
        for i in range(2, 10):
            work_sheet["A{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["J{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["D{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["G{}".format(i)].border = Border(right=border_design)
        list_index_columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
        for column in list_index_columns:
            work_sheet["{}1".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}4".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}7".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}10".format(column)].border = Border(bottom=border_design)
        # make modifications to show correctly
        work_sheet["B4"].border = Border(left=border_design, bottom=border_design)
        work_sheet["B7"].border = Border(left=border_design, bottom=border_design)
        work_sheet["B10"].border = Border(left=border_design, bottom=border_design)
        work_sheet["J4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["J7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["J10"].border = Border(right=border_design, bottom=border_design)

        work_sheet["D4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["D7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["D10"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G10"].border = Border(right=border_design, bottom=border_design)
        #  modifying cell style for cells that are completed
        color = openpyxl.styles.colors.Color(rgb="00F5FFDE")
        custom_fill = PatternFill(fill_type="solid", fgColor=color)
        list_digits = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        for idx in range(2, 11):
            for column_letter in list_index_columns:
                highlighted_cell = work_sheet["{}{}".format(column_letter, idx)]
                if highlighted_cell.value in list_digits:
                    work_sheet["{}{}".format(column_letter, idx)].fill = custom_fill
                    if (idx == 4 or idx == 7 or idx == 10) and (
                            column_letter == "D" or column_letter == "G" or column_letter == "J"):
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=border_design, right=border_design)
                    elif idx == 4 or idx == 7 or idx == 10:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"), left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"), bottom=border_design)
                    elif column_letter == "D" or column_letter == "G" or column_letter == "J":
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=Side(style='thin', color="FFC8C8C8"),
                            right=border_design)
                    elif column_letter == "B" and idx == 10:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=border_design, left=border_design)
                    else:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"), left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"), bottom=Side(style='thin', color="FFC8C8C8"))
        # correct b10
        if work_sheet["B10"].value in list_digits:
            work_sheet["A10"].border = Border(
                left=Side(style='thin', color="FFC8C8C8"), right=border_design,
                top=Side(style='thin', color="FFC8C8C8"), bottom=Side(style='thin', color="FFC8C8C8"))
        wb.save(constants.NAME_FILE)
        print("Initial game is loading....")
        time.sleep(2)
        os.system(constants.NAME_FILE)

    def write_solution_to_file(self, board_game):
        # copy what we have already in the initial board to a new Excel file
        file_path = constants.FILE_PATH
        os.chdir(file_path)
        wb = openpyxl.load_workbook(filename=constants.NAME_FILE)
        wb.create_sheet('solved game')

        max_row = wb['initial game'].max_row
        max_column = wb['initial game'].max_column
        for i in range(1, max_row+1):
            for j in range(1, max_column+1):
                wb['solved game'].cell(row=i, column=j).value = wb['initial game'].cell(row=i, column=j).value
        # re-style everything back
        wb.active = wb["solved game"]

        # copy every  value from the board to it's destination
        work_sheet = wb['solved game']
        border_design = Side(style="thick", color="FF0000")
        for i in range(2, 10):
            work_sheet["A{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["J{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["D{}".format(i)].border = Border(right=border_design)
        for i in range(1, 10):
            work_sheet["G{}".format(i)].border = Border(right=border_design)
        list_index_columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
        for column in list_index_columns:
            work_sheet["{}1".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}4".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}7".format(column)].border = Border(bottom=border_design)
        for column in list_index_columns:
            work_sheet["{}10".format(column)].border = Border(bottom=border_design)
        # make modifications to show correctly
        work_sheet["B4"].border = Border(left=border_design, bottom=border_design)
        work_sheet["B7"].border = Border(left=border_design, bottom=border_design)
        work_sheet["B10"].border = Border(left=border_design, bottom=border_design)
        work_sheet["J4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["J7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["J10"].border = Border(right=border_design, bottom=border_design)

        work_sheet["D4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["D7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["D10"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G4"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G7"].border = Border(right=border_design, bottom=border_design)
        work_sheet["G10"].border = Border(right=border_design, bottom=border_design)
        # modifying cell style for cells that are completed
        color = openpyxl.styles.colors.Color(rgb="00F5FFDE")
        custom_fill = PatternFill(fill_type="solid", fgColor=color)
        list_digits = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        for idx in range(2, 11):
            for column_letter in list_index_columns:
                highlighted_cell = work_sheet["{}{}".format(column_letter, idx)]
                if highlighted_cell.value in list_digits:
                    work_sheet["{}{}".format(column_letter, idx)].fill = custom_fill
                    if (idx == 4 or idx == 7 or idx == 10) and (
                            column_letter == "D" or column_letter == "G" or column_letter == "J"):
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=border_design, right=border_design)
                    elif idx == 4 or idx == 7 or idx == 10:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"), left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"), bottom=border_design)
                    elif column_letter == "D" or column_letter == "G" or column_letter == "J":
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=Side(style='thin', color="FFC8C8C8"),
                            right=border_design)
                    elif column_letter == "B" and idx == 10:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"),
                            bottom=border_design, left=border_design)
                    else:
                        work_sheet["{}{}".format(column_letter, idx)].border = Border(
                            right=Side(style='thin', color="FFC8C8C8"), left=Side(style='thin', color="FFC8C8C8"),
                            top=Side(style='thin', color="FFC8C8C8"), bottom=Side(style='thin', color="FFC8C8C8"))
        # correct b10
        if work_sheet["B10"].value in list_digits:
            work_sheet["A10"].border = Border(
                left=Side(style='thin', color="FFC8C8C8"), right=border_design,
                top=Side(style='thin', color="FFC8C8C8"), bottom=Side(style='thin', color="FFC8C8C8"))
        #introduce the values from the board
            # copy every  value from the board to it's destination
        for idx_row in range(2, 11):
            for idx_col in range(2, 11):
                highlighted_cell = work_sheet.cell(row=idx_row, column=idx_col)
                if highlighted_cell.value not in list_digits:
                    highlighted_cell.value = board_game[
                        constants.DICT_EXCEL_TO_BOARD[idx_row]][constants.DICT_EXCEL_TO_BOARD[idx_col]]
                    work_sheet.cell(row=idx_row, column=idx_col).font = Font(color="FF426ADD")
        wb.save(constants.NAME_FILE)
        print("Solution is being loaded...")
        time.sleep(3)
        os.system(constants.NAME_FILE)
